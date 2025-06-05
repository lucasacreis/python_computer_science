import openpyxl as px

while True:
    try:
        n = int(input('Digite o número do exercício: '))
        if 1 <= n <= 21:
            print(f'Você escolheu o exercício {n}.')
            break
        else:
            raise ValueError('Número do exercício deve estar entre 1 e 21.')
    except ValueError as e:
        print(e)

match n:
    case 10 | 11 | 12 | 13:
        # Mostrar folhas de uma planilha
        wb = px.load_workbook('exercices/data/medidas_satelite.xlsx')

        if n == 10:
            # Exibir o nome das folhas
            print(f'Folhas na planilha: {wb.sheetnames}')
        elif n == 11:
            # Exibir folha modelo de engenharia
            print(f'Folha modelo de engenharia: {wb.sheetnames[0]}')
            print(f'Folha modelo de engenharia: {wb['modelo_engenharia']}')
            print(f'Folha modelo de engenharia: {wb["modelo_engenharia"].title}')
        elif n == 12:
            # Exibir valores das celulas da folha modelo de engenharia
            print(f'Valores da folha modelo de engenharia:')
            print(f'{wb["modelo_engenharia"]["A1"].value}')
            print(f'{wb["modelo_engenharia"]["B1"].value}')
            print(f'{wb["modelo_engenharia"].cell(row=2, column=2).value}')
            print(f'Numero de linhas: {wb["modelo_engenharia"].max_row}')
            print(f'Numero de colunas: {wb["modelo_engenharia"].max_column}')

        wb.close()
